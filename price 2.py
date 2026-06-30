import os
import sys
import json
import base64
import time
import re
from io import BytesIO
from pathlib import Path
from pdf2image import convert_from_path
from openai import AzureOpenAI
import pandas as pd
from dotenv import load_dotenv
from openpyxl.styles import Font, PatternFill, Alignment
from ddgs import DDGS

SCRIPT_DIR = Path(__file__).resolve().parent
ENV_PATH = SCRIPT_DIR / ".env"
load_dotenv(dotenv_path=ENV_PATH)

# Load all Azure OpenAI configuration from .env file
AZURE_API_KEY = os.getenv("AZURE_OPENAI_API_KEY")
AZURE_ENDPOINT = os.getenv("AZURE_OPENAI_ENDPOINT")
AZURE_DEPLOYMENT = os.getenv("AZURE_OPENAI_DEPLOYMENT")
AZURE_API_VERSION = os.getenv("AZURE_OPENAI_API_VERSION", "2024-08-01-preview")

# Toggle DEBUG to see raw GPT-4o responses in terminal for troubleshooting
DEBUG = False

# Validate that all required Azure credentials are present
if not all([AZURE_API_KEY, AZURE_ENDPOINT, AZURE_DEPLOYMENT]):
    print("ERROR: Missing Azure OpenAI credentials in .env file.")
    sys.exit(1)

# Set poppler path from the bin folder next to this script
POPPLER_PATH = SCRIPT_DIR / "bin"
if not POPPLER_PATH.exists():
    print(f"ERROR: Poppler not found at {POPPLER_PATH}")
    sys.exit(1)

# Initialize Azure OpenAI client with endpoint and API version
client = AzureOpenAI(
    api_key=AZURE_API_KEY,
    azure_endpoint=AZURE_ENDPOINT,
    api_version=AZURE_API_VERSION
)

def get_folder_path_from_user():
    print("Construction Diagram Part Extractor & Cost Finder (Batch Mode with Live Search)")
    # Accept folder path as command-line argument or prompt user to enter it
    if len(sys.argv) > 1:
        folder_path = Path(sys.argv[1])
    else:
        user_input = input("\nEnter the full path to the folder containing your PDFs:\n> ").strip().strip('"').strip("'")
        folder_path = Path(user_input)
    if not folder_path.exists() or not folder_path.is_dir():
        print(f"ERROR: Folder not found: {folder_path}")
        sys.exit(1)
    # Collect all PDF files within the folder
    pdf_files = sorted(folder_path.glob("*.pdf"))
    if not pdf_files:
        print(f"ERROR: No PDF files found in {folder_path}")
        sys.exit(1)
    print(f"Found {len(pdf_files)} PDF file(s) in folder.")
    return folder_path, pdf_files

def encode_image_to_base64(image):
    # Convert PIL image to base64 JPEG string for GPT-4o vision input
    buffered = BytesIO()
    image.save(buffered, format="JPEG", quality=95)
    return base64.b64encode(buffered.getvalue()).decode("utf-8")

def split_image_into_tiles(image, overlap_pct=0.1):
    # Split a large page image into 2x2 overlapping tiles plus the full page for thorough text detection
    width, height = image.size
    overlap_w = int(width * overlap_pct)
    overlap_h = int(height * overlap_pct)
    mid_w = width // 2
    mid_h = height // 2
    # Define 4 overlapping quadrants so text near tile boundaries is not cut off
    tiles = [
        ("full_page", image),
        ("top_left", image.crop((0, 0, mid_w + overlap_w, mid_h + overlap_h))),
        ("top_right", image.crop((mid_w - overlap_w, 0, width, mid_h + overlap_h))),
        ("bottom_left", image.crop((0, mid_h - overlap_h, mid_w + overlap_w, height))),
        ("bottom_right", image.crop((mid_w - overlap_w, mid_h - overlap_h, width, height))),
    ]
    return tiles

def process_diagram_with_gpt4o(base64_image, label):
    prompt = """
You are analyzing a real engineering/construction diagram image (or a section of one).

STEP 1 - READ THE IMAGE CAREFULLY:
Read EVERY piece of text visible — including small text, callouts, leader lines, notes, title blocks, material specs, dimensions, machining notes, BOM tables, and tiny annotations. Be exhaustive. Engineering drawings often have 20-50 text items.

STEP 2 - CLASSIFY EACH TEXT:
For each text you read, decide if it is a valid part, component, material, or machining specification.

VALID (include these):
- Any drill/tap/ream/bore/counterbore callout with sizes (e.g. M10 TAP, M20 TAP, Ø10 DRILL, M39 TAP)
- Plates, sheets, rods, bars, profiles, angles, channels with material grade or thickness
- Fasteners (BOLT, NUT, SCREW, WASHER, STUD, RIVET) with sizes
- Bearings, gears, shafts, pulleys, sprockets with specs
- Valves, pumps, motors, cylinders, actuators with specs
- Welding callouts and surface finish notes describing manufacturing operations
- Any item listed in a parts list, BOM table, or material list
- Equipment or machine names in the title block

INVALID (ignore these):
- A number all by itself with no descriptive word
- A pure dimension with no descriptive word (just "50mm" or "1:100")
- Single isolated letters
- Drawing number, sheet number, revision letter, date, scale ratio

STRICT RULES:
- Extract ONLY text that genuinely appears in the image.
- Do NOT invent or guess any part names.
- Copy text EXACTLY as written.
- Be EXHAUSTIVE - extract every valid item you can see, even small ones.
- If you see "M10 TAP", "M20 TAP", "M39 TAP" each as separate callouts, list all of them.

STEP 3 - ESTIMATE PRICE:
For each valid item, estimate a realistic average market cost in Indian Rupees (INR ₹).

OUTPUT FORMAT:
Return ONLY a JSON array. No markdown. No code fences. No explanation.
Each item: {"Name": "exact text from image", "Cost": "₹amount"}
If you genuinely see zero valid items, return: []
"""
    try:
        # Use AZURE_DEPLOYMENT name as the model parameter for Azure OpenAI
        response = client.chat.completions.create(
            model=AZURE_DEPLOYMENT,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt},
                        {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{base64_image}", "detail": "high"}}
                    ],
                }
            ],
            max_tokens=4000,
            temperature=0.0
        )
        result_text = response.choices[0].message.content.strip()
        # Print raw GPT-4o output for debugging if DEBUG mode is enabled
        if DEBUG:
            print(f"    --- RAW ({label}) ---\n{result_text}\n    --- END ---")
        # Strip markdown code block formatting if GPT-4o adds it
        if result_text.startswith("```json"):
            result_text = result_text[7:].strip()
        if result_text.startswith("```"):
            result_text = result_text[3:].strip()
        if result_text.endswith("```"):
            result_text = result_text[:-3].strip()
        return json.loads(result_text)
    except json.JSONDecodeError:
        return []
    except Exception as e:
        print(f"      ERROR on tile {label}: {str(e)[:80]}")
        return []

def simplify_part_name_for_search(part_name):
    # Convert technical part text into a simpler search query for better web results
    prompt = f"""
Convert this engineering callout into a short product search query (max 5-6 words) for Indian e-commerce sites.

Engineering text: "{part_name}"

Rules:
- Extract the core item/material/operation
- Remove drawing-specific notes (hole counts, "Nos", "PCD", "DEEP")
- Keep important sizes/grades that affect price
- Output ONLY the search query

Examples:
"Ø10.2 DRILL M12 TAP 30 DEEP 3 Nos" -> M12 tap
"BASE PLATE 505.2X85X16 THK E250A IS:2062" -> MS plate IS 2062 E250A
"M10 25 DEEP 8 NOS" -> M10 tap

Your output:"""
    try:
        response = client.chat.completions.create(
            model=AZURE_DEPLOYMENT,
            messages=[{"role": "user", "content": prompt}],
            max_tokens=30,
            temperature=0.0
        )
        return response.choices[0].message.content.strip().replace('"', '').replace("'", "")
    except Exception:
        return part_name

def search_duckduckgo(query, max_results=6, retries=2):
    # Search DuckDuckGo for live price information about the part with retry logic
    for attempt in range(retries):
        try:
            with DDGS(timeout=15) as ddgs:
                results = list(ddgs.text(query, region="in-en", max_results=max_results))
            # Combine title and snippet from each result for GPT-4o context
            snippets = []
            for r in results:
                title = r.get("title", "")
                body = r.get("body", "")
                snippets.append(f"- {title}: {body}")
            return "\n".join(snippets) if snippets else ""
        except Exception:
            if attempt < retries - 1:
                time.sleep(2)
                continue
            return ""
    return ""

def parse_price_to_number(price_str):
    # Convert a price string like "₹1,500" or "₹2000" into an integer for comparison
    if not price_str:
        return None
    digits = re.sub(r'[^\d]', '', price_str)
    return int(digits) if digits else None

def refine_price_with_live_search(part_name, estimated_cost):
    # Use live web data to improve the estimated cost, but verify the price is sane before using it
    simplified_query = simplify_part_name_for_search(part_name)
    search_results = search_duckduckgo(f"{simplified_query} price India")
    if not search_results:
        return estimated_cost
    # Ask GPT-4o to read the search snippets and pick the most likely real price
    prompt = f"""
You are a price verification expert for Indian industrial/construction products.

Part name: "{part_name}"
Simplified search term used: "{simplified_query}"
GPT estimate: {estimated_cost}

Live web search results:
{search_results}

Task:
1. Look at the search results carefully.
2. Find prices in the snippets that are SPECIFICALLY for this exact part/material (not random numbers, model codes, or unrelated products).
3. If you find clear, relevant prices for this exact item: return the most realistic/median price in INR.
4. If the search results are unrelated, ambiguous, or have no clear price: respond with EXACTLY the word KEEP.
5. Never return outlier prices that are 10x more or 10x less than the GPT estimate unless you are very sure.

Respond with ONLY one of these:
- A price like: ₹XXX or ₹X,XXX
- Or the word: KEEP

Your response:"""
    try:
        response = client.chat.completions.create(
            model=AZURE_DEPLOYMENT,
            messages=[{"role": "user", "content": prompt}],
            max_tokens=30,
            temperature=0.0
        )
        answer = response.choices[0].message.content.strip().replace('"', '').replace("'", "")
        # If GPT-4o says KEEP it means search results were unreliable
        if "keep" in answer.lower() or "₹" not in answer:
            return estimated_cost
        # Sanity check: reject prices wildly different from the estimate
        live_value = parse_price_to_number(answer)
        est_value = parse_price_to_number(estimated_cost)
        if live_value and est_value:
            ratio = live_value / est_value if est_value > 0 else 0
            # Reject if live price is more than 10x or less than 1/10th the estimate
            if ratio > 10 or ratio < 0.1:
                return estimated_cost
        return answer
    except Exception:
        return estimated_cost

def remove_duplicates(parts_list):
    # Remove duplicate part entries by comparing lowercase names
    seen = set()
    unique_parts = []
    for part in parts_list:
        name_lower = part["Name"].strip().lower()
        if name_lower not in seen:
            seen.add(name_lower)
            unique_parts.append(part)
    return unique_parts

def process_single_pdf(pdf_path):
    # Convert one PDF into images and run multi-tile GPT-4o extraction for thorough coverage
    print(f"\n=== Processing: {pdf_path.name} ===")
    try:
        images = convert_from_path(str(pdf_path), dpi=300, poppler_path=str(POPPLER_PATH))
    except Exception as e:
        print(f"  ERROR: Failed to convert PDF.\nDetails: {e}")
        return []
    print(f"  Converted {len(images)} page(s).")
    all_parts = []
    # Step 1: extract parts from each page using full image plus 4 overlapping quadrants
    for i, image in enumerate(images):
        print(f"  Page {i + 1} of {len(images)} - splitting into tiles for thorough detection...")
        tiles = split_image_into_tiles(image)
        for tile_name, tile_image in tiles:
            base64_image = encode_image_to_base64(tile_image)
            parts = process_diagram_with_gpt4o(base64_image, label=f"page{i+1}_{tile_name}")
            if parts:
                print(f"    [{tile_name}] found {len(parts)} part(s).")
                all_parts.extend(parts)
    # Deduplicate across all tiles since the same part may appear in multiple tiles
    unique_parts = remove_duplicates(all_parts)
    print(f"  Total unique parts after deduplication: {len(unique_parts)}")
    # Step 2: refine each price by feeding live search results into GPT-4o with sanity checks
    refined_parts = []
    for part in unique_parts:
        print(f"    Refining price for: {part['Name']}")
        refined_price = refine_price_with_live_search(part["Name"], part["Cost"])
        print(f"      Estimate: {part['Cost']}  →  Final: {refined_price}")
        refined_parts.append({"Name": part["Name"], "Cost": refined_price})
        # Small delay to avoid being rate-limited
        time.sleep(1.5)
    return refined_parts

def save_combined_excel(results_by_pdf, folder_path):
    # Build a combined Excel sheet with PDF name headers separating each PDF's parts
    excel_path = folder_path / "All_PDFs_Parts_Cost.xlsx"
    rows = []
    for pdf_name, parts in results_by_pdf.items():
        # Add a header row containing the PDF name
        rows.append({"S.No": pdf_name, "Name": "", "Cost": ""})
        if parts:
            for idx, part in enumerate(parts, start=1):
                rows.append({"S.No": idx, "Name": part["Name"], "Cost": part["Cost"]})
        else:
            rows.append({"S.No": "-", "Name": "No valid parts found", "Cost": "-"})
        # Add a blank separator row between PDFs
        rows.append({"S.No": "", "Name": "", "Cost": ""})
    df = pd.DataFrame(rows, columns=["S.No", "Name", "Cost"])
    # Write to Excel and apply formatting to PDF name header rows
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Parts & Costs")
        worksheet = writer.sheets["Parts & Costs"]
        # Style each PDF name row with bold text and a light yellow background
        header_fill = PatternFill(start_color="FFEB99", end_color="FFEB99", fill_type="solid")
        header_font = Font(bold=True, size=12)
        for row_idx, row in enumerate(rows, start=2):
            if row["S.No"] in results_by_pdf:
                for col in range(1, 4):
                    cell = worksheet.cell(row=row_idx, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="left")
        # Auto-fit column widths based on the longest cell value in each column
        for column in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in column if cell.value)
            worksheet.column_dimensions[column[0].column_letter].width = max_length + 5
    return excel_path

def main():
    folder_path, pdf_files = get_folder_path_from_user()
    # Dictionary mapping each PDF filename to its extracted parts list
    results_by_pdf = {}
    for pdf_path in pdf_files:
        parts = process_single_pdf(pdf_path)
        results_by_pdf[pdf_path.name] = parts
    # Save the combined results for all PDFs into a single Excel file
    excel_path = save_combined_excel(results_by_pdf, folder_path)
    print(f"\nSUCCESS! Combined Excel saved to: {excel_path}")

if __name__ == "__main__":
    main()