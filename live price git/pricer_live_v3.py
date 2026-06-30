# SAVE AS: pricer_live_v3.py
# RUN AS: python pricer_live_v3.py material.png

import os, sys, json, base64, re, time, requests
from io import BytesIO
from pathlib import Path
from datetime import datetime

# dependency checks
try:
    from PIL import Image
except ImportError:
    sys.exit("Run: pip install Pillow")
try:
    from openai import AzureOpenAI
except ImportError:
    sys.exit("Run: pip install openai")
try:
    import pandas as pd
except ImportError:
    sys.exit("Run: pip install pandas openpyxl")
try:
    from dotenv import load_dotenv
except ImportError:
    sys.exit("Run: pip install python-dotenv")
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# startup banner
TODAY = datetime.now().strftime("%d %b %Y")
NOW_TIME = datetime.now().strftime("%H:%M")
print("=" * 65)
print(" PRICER LIVE-v3  —  Free Live Indian Market Prices")
print(f" Date : {TODAY}   Time : {NOW_TIME}")
print(" Sources: MCX scrape + SteelMint RSS + Moneycontrol +")
print("          Goodreturns + IndiaMART scrape + SerpAPI")
print("=" * 65)

# load credentials from .env
SCRIPT_DIR = Path(__file__).resolve().parent
load_dotenv(dotenv_path=SCRIPT_DIR / ".env")
API_KEY = os.getenv("AZURE_OPENAI_API_KEY")
ENDPOINT = os.getenv("AZURE_OPENAI_ENDPOINT")
DEPLOYMENT = os.getenv("AZURE_OPENAI_DEPLOYMENT")
API_VER = os.getenv("AZURE_OPENAI_API_VERSION", "2024-08-01-preview")
SERPAPI_KEY = os.getenv("SERPAPI_KEY", "")
if not all([API_KEY, ENDPOINT, DEPLOYMENT]):
    sys.exit("ERROR: Missing Azure OpenAI credentials in .env")

# init Azure OpenAI client
client = AzureOpenAI(api_key=API_KEY, azure_endpoint=ENDPOINT, api_version=API_VER)
print(f" Azure   : {DEPLOYMENT}")
print(f" SerpAPI : {'ENABLED' if SERPAPI_KEY else 'DISABLED — add SERPAPI_KEY to .env'}")
print("=" * 65)

# common HTTP headers for scraping
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept-Language": "en-IN,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

# safe HTTP GET wrapper - returns None on failure
def safe_get(url, params=None, timeout=12):
    try:
        r = requests.get(url, params=params, headers=HEADERS, timeout=timeout, verify=False)
        if r.status_code == 200 and len(r.text) > 200:
            return r
    except Exception:
        pass
    return None

# scrape MCX India commodity futures prices
def fetch_mcx_prices():
    results = {}
    commodities = {
        "steel": ["STEELF", "STEEL"], "aluminium": ["ALUMINIUM", "ALUMINIF"],
        "copper": ["COPPER", "COPPERM"], "zinc": ["ZINC", "ZINCM"],
        "lead": ["LEAD", "LEADM"], "nickel": ["NICKEL"],
    }
    per_quintal = {"steel"}
    base_url = "https://www.mcxindia.com/market-data/commodity-details"
    for metal, symbols in commodities.items():
        for sym in symbols:
            r = safe_get(f"{base_url}/{sym}")
            if not r: continue
            html = r.text
            patterns = [
                r'id="sp_ltp"[^>]*>([\d,\.]+)', r'LTP[^<]{0,50}([\d,]+\.?\d*)',
                r'Last\s*Price[^<]{0,50}([\d,]+\.?\d*)', r'"ltp"\s*:\s*"?([\d,\.]+)"?',
                r'class="[^"]*ltp[^"]*"[^>]*>([\d,\.]+)',
            ]
            for pat in patterns:
                m = re.search(pat, html, re.IGNORECASE)
                if m:
                    try:
                        val = float(m.group(1).replace(",", ""))
                        if metal in per_quintal and 20000 < val < 100000:
                            results[metal] = round(val / 100, 2); break
                        elif metal not in per_quintal and 50 < val < 200000:
                            results[metal] = round(val, 2); break
                    except ValueError: continue
                if metal in results: break
            if metal in results: break
    if results: print(f"  [MCX] Fetched: {results}")
    return results

# scrape Moneycontrol commodity prices
def fetch_moneycontrol_prices():
    results = {}
    urls = {
        "steel": "https://www.moneycontrol.com/commodity/steelf-futures-price.html",
        "aluminium": "https://www.moneycontrol.com/commodity/aluminium-futures-price.html",
        "copper": "https://www.moneycontrol.com/commodity/copper-futures-price.html",
        "zinc": "https://www.moneycontrol.com/commodity/zinc-futures-price.html",
        "lead": "https://www.moneycontrol.com/commodity/lead-futures-price.html",
        "nickel": "https://www.moneycontrol.com/commodity/nickel-futures-price.html",
    }
    per_quintal = {"steel"}
    for metal, url in urls.items():
        if metal in results: continue
        r = safe_get(url)
        if not r: continue
        html = r.text
        patterns = [
            r'class="[^"]*curr_price[^"]*"[^>]*>([\d,\.]+)',
            r'id="[^"]*cur_price[^"]*"[^>]*>([\d,\.]+)',
            r'"price"\s*:\s*"?([\d,\.]+)"?',
            r'<span[^>]*class="[^"]*price[^"]*"[^>]*>([\d,\.]+)',
            r'₹\s*([\d,]+\.?\d*)',
        ]
        for pat in patterns:
            m = re.search(pat, html, re.IGNORECASE)
            if m:
                try:
                    val = float(m.group(1).replace(",", ""))
                    if metal in per_quintal and 20000 < val < 100000:
                        results[metal] = round(val / 100, 2); break
                    elif metal not in per_quintal and 50 < val < 200000:
                        results[metal] = round(val, 2); break
                except ValueError: continue
    if results: print(f"  [Moneycontrol] Fetched: {results}")
    return results

# parse SteelMint RSS feed for daily Indian steel prices
def fetch_steelmint_prices():
    results = {}
    rss_urls = [
        "https://www.steelmint.com/feed/",
        "https://www.steelmint.com/rss/news.xml",
        "https://www.steelmint.com/rss/price.xml",
    ]
    for url in rss_urls:
        r = safe_get(url, timeout=10)
        if not r: continue
        items = re.findall(r'<item>(.*?)</item>', r.text, re.DOTALL)
        for item in items[:10]:
            raw = re.sub(r'<[^>]+>', ' ', item)
            for pat in [
                r'(?:Rs\.?|INR)\s*([\d,]+)\s*/\s*(?:MT|tonne|ton)',
                r'(?:Rs\.?|INR)\s*([\d,]+)\s*/\s*(?:quintal|qtl)',
                r'([\d,]+)\s*(?:Rs\.?|INR)\s*/\s*(?:MT|tonne|ton)',
            ]:
                for m in re.finditer(pat, raw, re.IGNORECASE):
                    try:
                        val = float(m.group(1).replace(",", ""))
                        ctx = raw[max(0, m.start()-50):m.end()+50].lower()
                        if ("steel" in ctx or "billet" in ctx or "hrc" in ctx) and 30000 < val < 120000:
                            results.setdefault("steel", round(val/1000, 2))
                        if "alumin" in ctx and 100000 < val < 400000:
                            results["aluminium"] = round(val/1000, 2)
                    except ValueError: continue
    if results: print(f"  [SteelMint] Fetched: {results}")
    return results

# scrape Goodreturns.in metal price pages
def fetch_goodreturns_prices():
    results = {}
    urls = {
        "steel": "https://www.goodreturns.in/commodity/steel-price-in-india.html",
        "aluminium": "https://www.goodreturns.in/commodity/aluminium-price-in-india.html",
        "copper": "https://www.goodreturns.in/commodity/copper-price-in-india.html",
    }
    for metal, url in urls.items():
        if metal in results: continue
        r = safe_get(url)
        if not r: continue
        html = r.text
        patterns = [
            r'(?:₹|Rs\.?)\s*([\d,]+\.?\d*)\s*(?:per\s*kg|/kg)',
            r'Today.*?(?:₹|Rs\.?)\s*([\d,]+\.?\d*)',
            r'Price.*?(?:₹|Rs\.?)\s*([\d,]+\.?\d*)',
            r'class="[^"]*price[^"]*"[^>]*>.*?(?:₹|Rs\.?)\s*([\d,]+\.?\d*)',
        ]
        for pat in patterns:
            m = re.search(pat, html, re.IGNORECASE | re.DOTALL)
            if m:
                try:
                    val = float(m.group(1).replace(",", ""))
                    if metal == "steel" and 30 < val < 200: results[metal] = round(val, 2); break
                    elif metal == "aluminium" and 100 < val < 500: results[metal] = round(val, 2); break
                    elif metal == "copper" and 400 < val < 2000: results[metal] = round(val, 2); break
                except ValueError: continue
    if results: print(f"  [Goodreturns] Fetched: {results}")
    return results

# SerpAPI call - real Google results with prices
def serpapi_live_price(query):
    if not SERPAPI_KEY: return "", ""
    try:
        r = requests.get("https://serpapi.com/search",
            params={"q": query, "gl": "in", "hl": "en", "num": "5", "api_key": SERPAPI_KEY}, timeout=20)
        data = r.json()
        for item in data.get("shopping_results", []):
            price, link = item.get("price", ""), item.get("link", "")
            if price and re.search(r'\d', price):
                price = re.sub(r'^(Rs\.?|INR)\s*', '₹', price)
                if not price.startswith("₹"): price = "₹" + price
                return price, link
        for item in data.get("organic_results", []):
            snippet, link = item.get("snippet", ""), item.get("link", "")
            m = re.search(
                r'(?:₹|Rs\.?|INR)\s*([\d,]+(?:\.\d+)?)(?:\s*[-–]\s*(?:₹|Rs\.?|INR)?\s*([\d,]+(?:\.\d+)?))?(?:\s*/\s*(?:kg|piece|pc|meter|mt|ton|unit))?',
                snippet, re.IGNORECASE)
            if m:
                raw = re.sub(r'^(Rs\.?|INR)\s*', '₹', m.group(0).strip())
                if not raw.startswith("₹"): raw = "₹" + raw
                return raw, link
        return "", ""
    except Exception as e:
        print(f"  [SerpAPI] {e}")
        return "", ""

# direct IndiaMART search page scrape
def indiamart_live_price(name, material):
    from urllib.parse import quote_plus
    q = quote_plus(f"{name} {material}")
    url = f"https://dir.indiamart.com/search.mp?ss={q}&prdsrc=1"
    r = safe_get(url)
    if not r: return "", ""
    prices = []
    pat = r'(?:₹|Rs\.?)\s*([\d,]+(?:\.\d+)?)(?:\s*[-–]\s*(?:₹|Rs\.?)?\s*([\d,]+(?:\.\d+)?))?(?:\s*/\s*(?:Kg|KG|kg|Piece|Nos|Unit|Meter|MT|Ton))?'
    for m in re.finditer(pat, r.text):
        try:
            low = float(m.group(1).replace(",", ""))
            if 1 < low < 500000: prices.append((low, m.group(0).strip()))
        except ValueError: continue
    if prices:
        prices.sort(key=lambda x: x[0])
        median_price = re.sub(r'^(Rs\.?)\s*', '₹', prices[len(prices)//2][1])
        if not median_price.startswith("₹"): median_price = "₹" + median_price
        return median_price, url
    return "", ""

# central cache - fetches live prices once at startup
class LivePriceCache:
    def __init__(self):
        self.prices, self.sources, self.fetched_at = {}, {}, ""
        self._fetch_all()
    def _fetch_all(self):
        print("\nFetching live metal prices (free sources)...")
        mcx, mc, sm, gr = fetch_mcx_prices(), fetch_moneycontrol_prices(), fetch_steelmint_prices(), fetch_goodreturns_prices()
        metals = ["steel", "aluminium", "copper", "zinc", "lead", "nickel"]
        for metal in metals:
            for src_dict, src_name in [(mcx, "MCX India"), (mc, "Moneycontrol"), (sm, "SteelMint"), (gr, "Goodreturns")]:
                if metal in src_dict and src_dict[metal] > 0:
                    self.prices[metal], self.sources[metal] = src_dict[metal], src_name
                    break
        self.fetched_at = datetime.now().strftime("%d %b %Y %H:%M")
        print(f"  Fetch complete: {self.fetched_at}")
        print(f"  Live prices: {self.prices}\n")
    def get(self, metal): return self.prices.get(metal.lower(), 0.0)
    def source(self, metal): return self.sources.get(metal.lower(), "static fallback")
    def retail_price(self, metal, markup_pct=12.0):
        spot = self.get(metal)
        return round(spot * (1 + markup_pct / 100), 2) if spot > 0 else 0.0
    def summary_text(self):
        lines = [f"Live metal prices as of {self.fetched_at}:"]
        markups = {"steel": 12, "aluminium": 14, "copper": 12, "zinc": 12, "lead": 12, "nickel": 14}
        for metal, spot in self.prices.items():
            mk = markups.get(metal, 12)
            lines.append(f"  {metal.capitalize():<12}: spot ₹{spot}/kg → retail ~₹{self.retail_price(metal, mk)}/kg (+{mk}%) [{self.sources.get(metal,'')}]")
        if not self.prices: lines.append("  No live data fetched — using static 2024 database")
        return "\n".join(lines)

# static IndiaMART/TradeIndia 2024 price database
DB_STATIC = [
    {"keys": ["base plate","top plate","bottom plate","cover plate","gusset plate","gusset","end plate","stiffener","ms plate","mild steel plate","steel plate","is:2062","is 2062","e250","e350","e250a","thk"],
     "price": "₹62-75/kg", "note": "MS Plate IS:2062 E250A 6-50mm",
     "url": "https://www.indiamart.com/proddetail/is-2062-e250a-ms-plate-2854558229355.html",
     "site": "IndiaMART — IS:2062 E250A MS Plate"},
    {"keys": ["round bar","ms rod","ms bar","rod","shaft","pin","spindle","axle"],
     "price": "₹58-68/kg", "note": "MS Round Bar IS:2062",
     "url": "https://www.indiamart.com/proddetail/ms-round-bar-22483223791.html",
     "site": "IndiaMART — MS Round Bar"},
    {"keys": ["flat bar","flat strip","ms flat"],
     "price": "₹60-70/kg", "note": "MS Flat Bar IS:2062",
     "url": "https://www.indiamart.com/proddetail/ms-flat-bar.html",
     "site": "IndiaMART — MS Flat Bar"},
    {"keys": ["ms pipe","erw pipe","pipe","tube","hollow section","rhs","shs"],
     "price": "₹65-78/kg", "note": "MS ERW Pipe IS:1239",
     "url": "https://www.indiamart.com/proddetail/ms-erw-pipe.html",
     "site": "IndiaMART — MS ERW Pipe"},
    {"keys": ["ismb","ismc","islb","isjb","isa","isfa","beam","channel","angle iron","joist","structural section"],
     "price": "₹62-74/kg", "note": "IS:2062 Structural Steel Sections",
     "url": "https://www.indiamart.com/proddetail/ismb-steel-beam.html",
     "site": "IndiaMART — IS:2062 Steel Sections"},
    {"keys": ["ss 304","ss304","stainless 304","aisi 304"],
     "price": "₹180-220/kg", "note": "SS 304 Stainless Steel Plate",
     "url": "https://www.indiamart.com/proddetail/stainless-steel-304-plate.html",
     "site": "IndiaMART — SS 304"},
    {"keys": ["ss 316","ss316","stainless 316","aisi 316"],
     "price": "₹220-270/kg", "note": "SS 316 Stainless Steel Plate",
     "url": "https://www.indiamart.com/proddetail/stainless-steel-316-plate.html",
     "site": "IndiaMART — SS 316"},
    {"keys": ["m16 bolt","m20 bolt","m24 bolt","anchor bolt","foundation bolt","ht bolt","high tensile bolt"],
     "price": "₹45-150/piece", "note": "M16-M24 HT Bolt Gr 8.8",
     "url": "https://www.indiamart.com/proddetail/m16-high-tensile-bolt.html",
     "site": "IndiaMART — M16-M24 HT Bolt"},
    {"keys": ["m6 bolt","m8 bolt","m10 bolt","m12 bolt","hex bolt"],
     "price": "₹12-35/piece", "note": "M6-M12 Hex Bolt Gr 8.8",
     "url": "https://www.indiamart.com/proddetail/hex-bolt-grade-8-8.html",
     "site": "IndiaMART — Hex Bolt"},
    {"keys": ["hex nut","nut","lock nut","nylock"],
     "price": "₹5-45/piece", "note": "Hex Nut IS:1364",
     "url": "https://www.indiamart.com/proddetail/hex-nut.html",
     "site": "IndiaMART — Hex Nut"},
    {"keys": ["washer","plain washer","spring washer","flat washer"],
     "price": "₹2-18/piece", "note": "MS Washer IS:2016",
     "url": "https://www.indiamart.com/proddetail/ms-washer.html",
     "site": "IndiaMART — Washer"},
    {"keys": ["bearing","ball bearing","roller bearing","deep groove","pillow block","skf","fag","6200","6300"],
     "price": "₹150-3500/piece", "note": "Deep Groove Ball Bearing",
     "url": "https://www.indiamart.com/proddetail/deep-groove-ball-bearing.html",
     "site": "IndiaMART — Ball Bearing"},
    {"keys": ["coupling","jaw coupling","flange coupling","flexible coupling"],
     "price": "₹500-9000/piece", "note": "Industrial Shaft Coupling",
     "url": "https://www.indiamart.com/proddetail/jaw-coupling.html",
     "site": "IndiaMART — Coupling"},
    {"keys": ["gearbox","gear box","reduction gearbox","speed reducer"],
     "price": "₹4000-40000/piece", "note": "Industrial Gearbox",
     "url": "https://www.indiamart.com/proddetail/industrial-gearbox.html",
     "site": "IndiaMART — Gearbox"},
    {"keys": ["motor","electric motor","induction motor","ac motor","3 phase motor","three phase motor"],
     "price": "₹4000-50000/piece", "note": "3-Phase AC Motor 1-15HP",
     "url": "https://www.indiamart.com/proddetail/three-phase-induction-motor.html",
     "site": "IndiaMART — AC Motor"},
    {"keys": ["pump","centrifugal pump","water pump","gear pump"],
     "price": "₹2500-45000/piece", "note": "Industrial Centrifugal Pump",
     "url": "https://www.indiamart.com/proddetail/centrifugal-pump.html",
     "site": "IndiaMART — Pump"},
    {"keys": ["valve","gate valve","globe valve","ball valve","butterfly valve"],
     "price": "₹400-9000/piece", "note": "Industrial Valve",
     "url": "https://www.indiamart.com/proddetail/gate-valve.html",
     "site": "IndiaMART — Valve"},
    {"keys": ["aluminium plate","aluminum plate","aluminium sheet","aluminum sheet","al plate","6061","6063"],
     "price": "₹185-235/kg", "note": "Aluminium Plate 6061/6063",
     "url": "https://www.indiamart.com/proddetail/aluminium-plate.html",
     "site": "IndiaMART — Aluminium Plate"},
    {"keys": ["aluminium rod","aluminum rod","al rod","al bar","aluminium bar"],
     "price": "₹175-220/kg", "note": "Aluminium Round Bar",
     "url": "https://www.indiamart.com/proddetail/aluminium-round-bar.html",
     "site": "IndiaMART — Aluminium Bar"},
    {"keys": ["cast iron","grey iron","ductile iron","sg iron"],
     "price": "₹48-68/kg", "note": "Grey Cast Iron",
     "url": "https://www.indiamart.com/proddetail/cast-iron-plate.html",
     "site": "IndiaMART — Cast Iron"},
    {"keys": ["copper plate","copper rod","copper bar","etp copper"],
     "price": "₹680-790/kg", "note": "Copper Plate/Rod",
     "url": "https://www.indiamart.com/proddetail/copper-plate.html",
     "site": "IndiaMART — Copper"},
    {"keys": ["brass plate","brass rod","brass bar","brass tube"],
     "price": "₹380-490/kg", "note": "Brass Rod/Plate",
     "url": "https://www.indiamart.com/proddetail/brass-rod.html",
     "site": "IndiaMART — Brass"},
    {"keys": ["nylon","pa6","pa66","polyamide"],
     "price": "₹180-270/kg", "note": "Nylon 6/PA66 Sheet/Rod",
     "url": "https://www.indiamart.com/proddetail/nylon-sheet.html",
     "site": "IndiaMART — Nylon"},
    {"keys": ["ptfe","teflon"],
     "price": "₹1200-2200/kg", "note": "PTFE/Teflon Sheet",
     "url": "https://www.indiamart.com/proddetail/ptfe-sheet.html",
     "site": "IndiaMART — PTFE"},
    {"keys": ["hdpe","high density polyethylene","hdpe pipe","hdpe sheet"],
     "price": "₹90-135/kg", "note": "HDPE Pipe/Sheet",
     "url": "https://www.indiamart.com/proddetail/hdpe-pipe.html",
     "site": "IndiaMART — HDPE"},
    {"keys": ["pvc","polyvinyl chloride","pvc pipe","pvc sheet"],
     "price": "₹75-115/kg", "note": "PVC Pipe/Sheet",
     "url": "https://www.indiamart.com/proddetail/pvc-pipe.html",
     "site": "IndiaMART — PVC"},
    {"keys": ["rubber sheet","rubber gasket","rubber pad","neoprene","epdm"],
     "price": "₹120-260/kg", "note": "Industrial Rubber Sheet",
     "url": "https://www.indiamart.com/proddetail/rubber-sheet.html",
     "site": "IndiaMART — Rubber"},
    {"keys": ["cement","opc","ppc","concrete","m20","m25","m30"],
     "price": "₹370-440/bag (50kg)", "note": "OPC 53 Grade Cement",
     "url": "https://www.indiamart.com/proddetail/opc-53-cement.html",
     "site": "IndiaMART — OPC Cement"},
    {"keys": ["plywood","marine ply","shuttering ply"],
     "price": "₹45-130/sq ft", "note": "Plywood IS:710",
     "url": "https://www.indiamart.com/proddetail/plywood.html",
     "site": "IndiaMART — Plywood"},
    {"keys": ["gi pipe","galvanized pipe","galvanised pipe","gi sheet"],
     "price": "₹72-90/kg", "note": "GI Pipe/Sheet IS:1239",
     "url": "https://www.indiamart.com/proddetail/gi-pipe.html",
     "site": "IndiaMART — GI"},
    {"keys": ["spring","compression spring","tension spring","helical spring"],
     "price": "₹50-600/piece", "note": "Industrial Spring",
     "url": "https://www.indiamart.com/proddetail/compression-spring.html",
     "site": "IndiaMART — Spring"},
    {"keys": ["pulley","v-belt pulley","flat belt pulley","sheave"],
     "price": "₹250-4000/piece", "note": "MS/CI V-Belt Pulley",
     "url": "https://www.indiamart.com/proddetail/v-belt-pulley.html",
     "site": "IndiaMART — Pulley"},
    {"keys": ["v-belt","v belt","vbelt","a section belt","b section belt"],
     "price": "₹80-600/piece", "note": "Industrial V-Belt",
     "url": "https://www.indiamart.com/proddetail/v-belt.html",
     "site": "IndiaMART — V-Belt"},
    {"keys": ["chain","roller chain","drive chain","conveyor chain"],
     "price": "₹800-6000/meter", "note": "Industrial Roller Chain",
     "url": "https://www.indiamart.com/proddetail/roller-chain.html",
     "site": "IndiaMART — Roller Chain"},
    {"keys": ["sprocket","chain sprocket","drive sprocket"],
     "price": "₹350-3500/piece", "note": "MS/CI Chain Sprocket",
     "url": "https://www.indiamart.com/proddetail/chain-sprocket.html",
     "site": "IndiaMART — Sprocket"},
    {"keys": ["gasket","flange gasket","spiral wound gasket"],
     "price": "₹50-800/piece", "note": "Industrial Flange Gasket",
     "url": "https://www.indiamart.com/proddetail/flange-gasket.html",
     "site": "IndiaMART — Gasket"},
    {"keys": ["key","keyway","parallel key","woodruff key","machine key"],
     "price": "₹25-180/piece", "note": "MS/EN8 Machine Key IS:2048",
     "url": "https://www.indiamart.com/proddetail/parallel-key.html",
     "site": "IndiaMART — Machine Key"},
]

# fallback prices when no DB keyword matches
MAT_FALLBACK = {
    "Steel": ("₹62-75/kg", "https://www.indiamart.com/proddetail/is-2062-mild-steel-plate.html", "IndiaMART — IS:2062 MS Plate"),
    "Cast Iron": ("₹48-68/kg", "https://www.indiamart.com/proddetail/cast-iron-plate.html", "IndiaMART — Cast Iron"),
    "Aluminium": ("₹185-235/kg", "https://www.indiamart.com/proddetail/aluminium-plate.html", "IndiaMART — Aluminium Plate"),
    "Copper/Brass": ("₹380-790/kg", "https://www.indiamart.com/proddetail/copper-plate.html", "IndiaMART — Copper/Brass"),
    "Plastic/Polymer": ("₹80-210/kg", "https://www.indiamart.com/proddetail/engineering-plastic-sheet.html", "IndiaMART — Engineering Plastic"),
    "Concrete/Cement": ("₹370-440/bag", "https://www.indiamart.com/proddetail/opc-53-cement.html", "IndiaMART — Cement"),
    "Wood/Timber": ("₹45-130/sq ft", "https://www.indiamart.com/proddetail/plywood.html", "IndiaMART — Plywood/Timber"),
    "Rubber": ("₹120-260/kg", "https://www.indiamart.com/proddetail/rubber-sheet.html", "IndiaMART — Rubber Sheet"),
    "Composite": ("₹250-620/kg", "https://www.indiamart.com/proddetail/frp-sheet.html", "IndiaMART — FRP Composite"),
    "Other": ("₹60-150/kg", "https://www.indiamart.com/", "IndiaMART"),
}

# material classification keywords
MAT_KEYWORDS = {
    "Steel": ["steel","ms ","m.s.","mild steel","is:2062","is 2062","e250","e350","e250a","stainless","ss ","galvanized","gi ","thk","ismb","ismc","islb","plate","flat bar","round bar"],
    "Cast Iron": ["cast iron","ci ","grey iron","ductile iron","sg iron"],
    "Aluminium": ["aluminium","aluminum","al ","6061","6063"],
    "Copper/Brass": ["copper","brass","bronze","cu "],
    "Plastic/Polymer": ["plastic","nylon","ptfe","pvc","hdpe","polymer","teflon"],
    "Concrete/Cement": ["concrete","cement","rcc","pcc","mortar","m20","m25","m30"],
    "Wood/Timber": ["timber","wood","plywood","mdf","teak"],
    "Rubber": ["rubber","neoprene","epdm","gasket","o-ring"],
    "Composite": ["frp","grp","fiberglass","carbon fibre"],
}

# maps Material Type to LME metal name for live spot pricing
MAT_TO_METAL = {"Steel": "steel", "Aluminium": "aluminium", "Copper/Brass": "copper", "Cast Iron": "steel"}

# remove markdown code fences from GPT response
def strip_json(t):
    t = t.strip()
    t = re.sub(r"^```(?:json)?\s*", "", t)
    t = re.sub(r"\s*```$", "", t).strip()
    return t

# convert image file to base64 JPEG for GPT-4o Vision
def encode_image(path):
    with Image.open(path) as img:
        if img.mode not in ("RGB", "L"): img = img.convert("RGB")
        buf = BytesIO()
        img.save(buf, format="JPEG", quality=95)
        return base64.b64encode(buf.getvalue()).decode()

# read image path from argv or prompt
def get_path():
    if len(sys.argv) > 1:
        p = Path(sys.argv[1])
    else:
        p = Path(input("\nImage path:\n> ").strip().strip('"').strip("'"))
    if not p.is_file(): sys.exit(f"File not found: {p}")
    return p

# find best DB entry by counting keyword matches
def db_lookup(name, desc):
    combined = (name + " " + desc).lower()
    best, best_n = None, 0
    for entry in DB_STATIC:
        n = sum(1 for k in entry["keys"] if k in combined)
        if n > best_n: best_n, best = n, entry
    return best if best_n >= 1 else None

# classify item by keyword, fall back to GPT
def classify(name, desc):
    combined = (name + " " + desc).lower()
    for mat, kws in MAT_KEYWORDS.items():
        if any(k in combined for k in kws): return mat
    try:
        resp = client.chat.completions.create(model=DEPLOYMENT,
            messages=[{"role":"user","content":
                f'One word — material category:\nSteel|Cast Iron|Aluminium|Copper/Brass|Plastic/Polymer|Concrete/Cement|Wood/Timber|Rubber|Composite|Other\nItem:"{name}" Desc:"{desc}"\nPlates/brackets/beams with no stated material = Steel'}],
            max_tokens=10, temperature=0.0)
        return resp.choices[0].message.content.strip()
    except Exception:
        return "Steel"

# normalise IS code formatting (e.g. "is 2062 e250a" → "IS:2062 E250A")
def normalize_is_code(raw):
    if not raw or not raw.strip(): return ""
    s = raw.strip()
    s = re.sub(r'\b[Ii][Ss][\s\-:]*(\d+)', r'IS:\1', s)
    s = re.sub(r'\bastm\s+', 'ASTM ', s, flags=re.IGNORECASE)
    s = re.sub(r'\b([eE])(\d+)([aAbB]?)\b',
               lambda m: m.group(1).upper() + m.group(2) + m.group(3).upper(), s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

# infer IS code from item name/description when not provided by GPT
def infer_is_code(name, desc, material):
    combined = (name + " " + desc).lower()
    m = re.search(r'\b[Ii][Ss][\s\-:]*\d+[a-zA-Z0-9 \-]*', name + " " + desc)
    if m: return normalize_is_code(m.group(0))
    default_codes = {
        ("plate","beam","section","ismb","ismc","isa","gusset","bracket","base plate","top plate","flat bar","round bar"): "IS:2062",
        ("pipe","tube","erw","hollow section","rhs","shs"): "IS:1239",
        ("bolt","nut","fastener","stud","ht bolt","anchor bolt"): "IS:1364",
        ("washer",): "IS:2016",
        ("cement","opc","ppc","concrete"): "IS:269",
        ("plywood","ply"): "IS:710",
        ("hdpe",): "IS:4984",
        ("pvc",): "IS:4985",
        ("aluminium","aluminum","al plate","al rod"): "IS:737",
        ("key","keyway","parallel key"): "IS:2048",
        ("motor","induction motor","3 phase motor"): "IS:325",
        ("gi ","galvanized","galvanised"): "IS:1239",
        ("brass",): "IS:319",
    }
    for keys, code in default_codes.items():
        if any(k in combined for k in keys): return code
    return ""

# call GPT-4o Vision to extract parts + IS codes from diagram image
def extract_parts(b64):
    prompt = """
Analyze this engineering diagram. Extract every purchasable physical item.

INCLUDE: structural parts, machines, fasteners, materials with specs, steel sections.
EXCLUDE: bare dimension numbers, drawing codes, revision notes.

IS CODE / MATERIAL GRADE EXTRACTION:
Look carefully in the MATERIAL column of any parts table, or anywhere in the image,
for IS codes and material grades. Common patterns:
  IS:2062, IS 2062, E250A IS:2062, E350 IS:2062, IS:1239, IS:1364, IS:2016,
  IS:269, ASTM A36, DIN 17100, EN 10025, SA516 Gr 70

If an IS code appears in the MATERIAL column of a parts table, it applies to
EVERY part in that table — extract it for each part.
If a single material spec covers multiple rows, assign the same IS code to all those parts.
If no IS code is visible, set "IS_Code" to "" (empty string).

Return ONLY a raw JSON array (no markdown):
[{"Name":"exact text","Description":"plain english + grade","IS_Code":"IS:2062 E250A"}]
If nothing found: []
"""
    resp = client.chat.completions.create(model=DEPLOYMENT,
        messages=[{"role":"user","content":[
            {"type":"text","text":prompt},
            {"type":"image_url","image_url":{"url":f"data:image/jpeg;base64,{b64}","detail":"high"}}
        ]}], max_tokens=4000, temperature=0.0)
    try:
        data = json.loads(strip_json(resp.choices[0].message.content))
        return data if isinstance(data, list) else []
    except Exception as e:
        print(f"  Parse error: {e}")
        return []

# main pricing engine - uses live web + spot + GPT reasoning
def gpt_price(name, desc, mat, db_hit, live):
    web_price, web_url = "", ""
    if SERPAPI_KEY:
        q = f"{name} {mat} price per kg India indiamart site:indiamart.com"
        web_price, web_url = serpapi_live_price(q)
        if web_price: print(f"        SerpAPI: {web_price}")
    if not web_price:
        web_price, web_url = indiamart_live_price(name, mat)
        if web_price: print(f"        IndiaMART scrape: {web_price}")
    metal = MAT_TO_METAL.get(mat, "")
    spot = live.get(metal) if metal else 0.0
    retail = live.retail_price(metal) if metal else 0.0
    spot_src = live.source(metal) if metal else ""
    live_summary = live.summary_text()
    web_block = (f"LIVE WEB PRICE (scraped today {TODAY}):\n  Price : {web_price}\n  URL   : {web_url}\n  → Use this as primary price."
        if web_price else "No live web price found — use spot/static data below.")
    spot_block = (f"LIVE SPOT PRICE for {metal.upper()} (from {spot_src}):\n  Spot  : ₹{spot}/kg\n  Retail: ₹{retail}/kg (spot + ~12% margin)\n  → Use retail price if no web price available."
        if spot > 0 else "No live spot price available for this material.")
    db_block = (f"STATIC DB MATCH: {db_hit['price']} — {db_hit['note']} ({db_hit['site']})"
        if db_hit else "No static DB match — rely on live data or table below.")
    prompt = f"""
You are a senior Indian procurement engineer. Today is {TODAY}.

ITEM:
  Name : {name}
  Desc : {desc}
  Mat  : {mat}

{web_block}

{spot_block}

{db_block}

{live_summary}

STATIC REFERENCE (use only if all live data unavailable):
  MS Plate IS:2062 E250A (6-50mm)    → 62-75 ₹/kg
  MS Round Bar IS:2062               → 58-68 ₹/kg
  Structural Sections ISMB/ISMC      → 62-74 ₹/kg
  SS 304 Plate                       → 180-220 ₹/kg
  Aluminium Plate 6061               → 185-235 ₹/kg
  Cast Iron                          → 48-68 ₹/kg
  Copper                             → 680-790 ₹/kg
  M16-M24 HT Bolt Gr8.8              → 45-150 ₹/piece
  Ball Bearing                       → 150-3500 ₹/piece
  AC Motor 1-15HP                    → 4000-50000 ₹/piece
  Centrifugal Pump                   → 2500-45000 ₹/piece

PRICING RULES:
1. If LIVE WEB PRICE shown → use it directly (highest priority)
2. If LIVE SPOT PRICE shown → use retail figure (spot + margin)
3. Otherwise use static DB match or table
4. Fabricated/machined parts: raw material price × 1.3 to 1.5
5. State price_type as LIVE if from web/spot, ESTIMATED if from static

MANDATORY:
- Always output a numeric price. Never say N/A.
- Use ₹ symbol
- Provide IndiaMART URL

Respond ONLY in this JSON (raw, no markdown):
{{"price":"₹XX-YY/unit","price_type":"LIVE|ESTIMATED","source_url":"https://...","source_name":"site","reasoning":"one sentence"}}
"""
    try:
        resp = client.chat.completions.create(model=DEPLOYMENT,
            messages=[{"role":"user","content":prompt}], max_tokens=300, temperature=0.0)
        parsed = json.loads(strip_json(resp.choices[0].message.content))
        price = str(parsed.get("price","")).strip()
        price_type = str(parsed.get("price_type","ESTIMATED")).strip()
        url = str(parsed.get("source_url","")).strip()
        site = str(parsed.get("source_name","")).strip()
        reason = str(parsed.get("reasoning","")).strip()
        print(f"        [{price_type}] {reason}")
        if not price or not re.search(r'\d', price): raise ValueError(f"Bad price: {price!r}")
        price = re.sub(r'^(Rs\.?\s*|INR\s*)', '₹', price)
        if not price.startswith("₹"): price = "₹" + price
        tag = f" (live {TODAY})" if price_type == "LIVE" else f" (est. {TODAY})"
        price += tag
        source = url or "https://www.indiamart.com/"
        if site: source += f"  [{site}]"
        return price, source
    except Exception as e:
        print(f"        GPT error: {e}")
    if web_price: return web_price + f" (live {TODAY})", web_url + f"  [Live IndiaMART {TODAY}]"
    if retail > 0: return f"₹{retail}/kg (live {TODAY})", f"Live spot via {spot_src}"
    if db_hit:
        p = db_hit["price"]
        if not p.startswith("₹"): p = "₹" + p
        return p + f" (est. {TODAY})", f"{db_hit['url']}  [{db_hit['site']}]"
    p, u, s = MAT_FALLBACK.get(mat, MAT_FALLBACK["Other"])
    return p + f" (est. {TODAY})", f"{u}  [{s}]"

# remove duplicate parts by name
def dedup(parts):
    seen, out = set(), []
    for p in parts:
        k = re.sub(r'\s+', ' ', p.get("Name", "")).strip().lower()
        if k and k not in seen:
            seen.add(k); out.append(p)
    return out

# write formatted Excel with IS Code column + Live Price Sources sheet
def save_excel(rows, out_path, live):
    df = pd.DataFrame(rows, columns=["S.No","Name","IS Code","Material Type","Cost","Source"])
    with pd.ExcelWriter(out_path, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name="Parts & Costs")
        ws = w.sheets["Parts & Costs"]
        hf = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        hb = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        ef = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        wf = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        nf = Font(name="Calibri", size=10)
        livef = Font(name="Calibri", size=10, bold=True, color="1A7A1A")
        iscf = Font(name="Calibri", size=10, bold=True, color="0F3B6E")
        for cell in ws[1]:
            cell.fill = hb; cell.font = hf
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 22
        for ri in range(2, len(rows) + 2):
            cost_val = str(ws.cell(row=ri, column=5).value or "")
            is_live = "(live" in cost_val.lower()
            for ci in range(1, 7):
                c = ws.cell(row=ri, column=ci)
                c.fill = ef if ri % 2 == 0 else wf
                if is_live and ci == 5: c.font = livef
                elif ci == 3: c.font = iscf
                else: c.font = nf
                c.alignment = Alignment(horizontal="center" if ci in (1, 3) else "left",
                                        vertical="center", wrap_text=ci in (2, 6))
            ws.row_dimensions[ri].height = 32
        for i, ww in enumerate([7, 38, 22, 20, 28, 72], start=1):
            ws.column_dimensions[get_column_letter(i)].width = ww
        ws.freeze_panes = "A2"
        # second sheet - documents live price sources
        ms = w.book.create_sheet("Live Price Sources")
        norm = Font(name="Calibri", size=10)
        ms["A1"] = f"Live Market Price Report — {TODAY}"
        ms["A1"].font = Font(bold=True, name="Calibri", size=12, color="1F4E79")
        rows_meta = [("Generated on", TODAY), ("Time", NOW_TIME), ("", ""),
                     ("=== LIVE METAL SPOT PRICES ===", "")]
        for metal, spot in live.prices.items():
            mk = 12
            rows_meta.append((f"  {metal.capitalize()}",
                f"Spot ₹{spot}/kg  →  Retail ~₹{live.retail_price(metal, mk)}/kg  (+{mk}% margin)  [{live.sources.get(metal,'')}]"))
        rows_meta += [("", ""), ("=== FREE PRICE SOURCES USED ===", ""),
            ("MCX India", "https://www.mcxindia.com"),
            ("Moneycontrol", "https://www.moneycontrol.com/commodity"),
            ("SteelMint RSS", "https://www.steelmint.com/feed"),
            ("Goodreturns.in", "https://www.goodreturns.in/commodity"),
            ("IndiaMART scrape", "https://dir.indiamart.com/search.mp"),
            ("SerpAPI", f"https://serpapi.com  [{'ENABLED' if SERPAPI_KEY else 'DISABLED'}]"),
            ("", ""), ("=== LEGEND ===", ""),
            ("Green bold (Cost)", "Live price fetched today"),
            ("Navy bold (IS Code)", "Indian Standard code from diagram or inferred"),
            ("Normal text", "Static reference / estimated")]
        for i, (k, v) in enumerate(rows_meta, start=3):
            ms.cell(row=i, column=1, value=k).font = (Font(bold=True, name="Calibri", size=10) if k.startswith("===") else norm)
            ms.cell(row=i, column=2, value=v).font = norm
        ms.column_dimensions["A"].width = 30
        ms.column_dimensions["B"].width = 70
    print(f"  Saved → {out_path}")

# orchestrator - runs the full pipeline
def main():
    path = get_path()
    print(f"\nImage : {path.name}\n")
    live = LivePriceCache()
    print(live.summary_text(), "\n")
    print("[1/4] Encoding image...")
    b64 = encode_image(path)
    print("[2/4] Extracting parts via GPT-4o Vision...")
    parts = dedup(extract_parts(b64))
    if not parts: sys.exit("No parts detected.")
    print(f"  → {len(parts)} part(s) found\n")
    print("[3/4] Pricing each part...\n")
    rows = []
    for i, p in enumerate(parts, 1):
        name = p.get("Name", "").strip()
        desc = p.get("Description", "").strip()
        is_code = normalize_is_code(p.get("IS_Code", "").strip())
        print(f"  [{i}/{len(parts)}] {name}")
        print(f"    Desc : {desc[:90]}")
        mat = classify(name, desc)
        if not is_code: is_code = infer_is_code(name, desc, mat)
        print(f"    IS   : {is_code or '(none)'}")
        full_desc = f"{desc} {is_code}".strip()
        hit = db_lookup(name, full_desc)
        cost, src = gpt_price(name, full_desc, mat, hit, live)
        print(f"    → {mat} | {cost}\n")
        rows.append({"S.No": i, "Name": name, "IS Code": is_code or "—",
                     "Material Type": mat, "Cost": cost, "Source": src})
        time.sleep(0.2)
    print("[4/4] Saving Excel...")
    out = path.parent / (path.stem + "_parts_cost_LIVE.xlsx")
    save_excel(rows, out, live)
    live_count = sum(1 for r in rows if "(live" in r["Cost"].lower())
    print(f"\n{'='*55}\n Done — {len(rows)} parts priced")
    print(f" Live prices : {live_count} items (green in Excel)")
    print(f" Estimated   : {len(rows)-live_count} items (static 2024 reference)")
    print(f" Output      : {out}\n{'='*55}")

if __name__ == "__main__":
    main()